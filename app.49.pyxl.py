import openpyxl as xl
from Tools.scripts.make_ctype import values
from openpyxl import cell
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    # cell = sheet['a1']
    # cell = sheet.cell(1, 1)
    # print(cell.value)  # name of the 1 to 1 cell on excel file
    # print(sheet.max_row)  # how many rows u can see on excel file

    # for row in range(1, sheet.max_row + 1):
    # print(row)  # we get the number of rows

    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, 3)
    print(cell.value)  # we got the values at third cell
    corrected_price = cell.value * 0.9  # we fixed the mistake at the cell
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

    Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')
    wb.save('filename')
